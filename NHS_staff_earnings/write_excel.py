import pandas as pd
from openpyxl import load_workbook

# Create function to be able to write to Excel for CORE organisations
def write_data_to_excel_CORE(dataframes_CORE, excel_template_CORE, excel_output_CORE, sheet_name):
    xl_writer = pd.ExcelWriter(excel_output_CORE, engine='openpyxl')
    wb = load_workbook(excel_template_CORE)
    xl_writer.book = wb
    xl_writer.sheets = {ws.title: ws for ws in wb.worksheets}

    for df, sheet_name in zip (dataframes_CORE, sheet_name):
        df.to_excel(xl_writer, sheet_name, index = False)
    xl_writer.save()

# Create function to be able to write to Excel for WIDER organisations
def write_data_to_excel_WIDER(dataframes_WIDER, excel_template_WIDER, excel_output_WIDER, sheet_name):
    xl_writer = pd.ExcelWriter(excel_output_WIDER, engine='openpyxl')
    wb = load_workbook(excel_template_WIDER)
    xl_writer.book = wb
    xl_writer.sheets = {ws.title: ws for ws in wb.worksheets}

    for df, sheet_name in zip (dataframes_WIDER, sheet_name):
        df.to_excel(xl_writer, sheet_name, index = False)
    xl_writer.save()